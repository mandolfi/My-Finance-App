from fastapi import FastAPI, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, or_
from database import get_db
from models import Account, Entrata, Uscita, Transazione
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = FastAPI(title="Patrimonio API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Filtro standard: solo transazioni reali (Prev vuoto o "chiusure")
def filtro_reali(query):
    return query.filter(or_(Transazione.prev.is_(None), Transazione.prev == "chiusure"))


# ── ACCOUNTS ────────────────────────────────────────────────────────────────

@app.get("/accounts")
def get_accounts(db: Session = Depends(get_db)):
    accounts = db.query(Account).order_by(Account.categoria, Account.nome_conto).all()
    return [{"id": a.id, "nome": a.nome_conto, "categoria": a.categoria} for a in accounts]


@app.get("/accounts/saldi")
def get_saldi(db: Session = Depends(get_db)):
    DATA_INIZIO_SALDI = date(2023, 1, 1)
    accounts = db.query(Account).order_by(Account.categoria, Account.nome_conto).all()
    risultati = []

    for a in accounts:
        query = filtro_reali(db.query(func.sum(Transazione.importo)).filter(
            Transazione.account_id == a.id,
            Transazione.data >= DATA_INIZIO_SALDI,
        ))
        saldo = query.scalar() or 0

        risultati.append({
            "id": a.id,
            "nome": a.nome_conto,
            "categoria": a.categoria,
            "saldo": round(float(saldo), 2),
        })

    patrimonio_liquido = sum(
        r["saldo"] for r in risultati
        if r["categoria"] not in ("debito", "carta di credito", "immobili")
    )
    valore_immobili = sum(
        r["saldo"] for r in risultati if r["categoria"] == "immobili"
    )
    debiti = sum(
        r["saldo"] for r in risultati
        if r["categoria"] in ("debito", "carta di credito")
    )
    patrimonio_totale = patrimonio_liquido + valore_immobili + debiti

    return {
        "conti": risultati,
        "patrimonio_liquido": round(patrimonio_liquido, 2),
        "valore_immobili": round(valore_immobili, 2),
        "totale_debiti": round(debiti, 2),
        "patrimonio_netto_totale": round(patrimonio_totale, 2),
    }


# ── DASHBOARD SUMMARY ───────────────────────────────────────────────────────

@app.get("/dashboard/summary")
def get_summary(anno: int = None, mese: int = None, db: Session = Depends(get_db)):
    if anno is None or mese is None:
        ultima = filtro_reali(db.query(Transazione)).order_by(Transazione.data.desc()).first()
        riferimento = ultima.data if ultima else date.today()
        anno = anno or riferimento.year
        mese = mese or riferimento.month

    base = filtro_reali(db.query(Transazione).filter(
        extract("year", Transazione.data) == anno,
        extract("month", Transazione.data) == mese,
    ))

    totale_entrate = base.filter(Transazione.direzione == "Entrata")\
        .with_entities(func.sum(Transazione.importo)).scalar() or 0
    totale_uscite_raw = base.filter(Transazione.direzione == "Uscita")\
        .with_entities(func.sum(Transazione.importo)).scalar() or 0

    # Le uscite in DB hanno segno negativo; per leggibilità mostriamo un numero positivo
    totale_uscite = abs(float(totale_uscite_raw))
    risparmio = float(totale_entrate) - totale_uscite
    tasso_risparmio = (risparmio / float(totale_entrate) * 100) if totale_entrate else 0

    return {
        "anno": anno,
        "mese": mese,
        "totale_entrate": round(float(totale_entrate), 2),
        "totale_uscite": round(totale_uscite, 2),
        "risparmio": round(risparmio, 2),
        "tasso_risparmio": round(tasso_risparmio, 1),
    }


# ── CASHFLOW ULTIMI MESI ─────────────────────────────────────────────────────

@app.get("/dashboard/cashflow")
def get_cashflow(db: Session = Depends(get_db)):
    query = filtro_reali(db.query(
        extract("year", Transazione.data).label("anno"),
        extract("month", Transazione.data).label("mese"),
        Transazione.direzione,
        func.sum(Transazione.importo).label("totale"),
    )).group_by("anno", "mese", Transazione.direzione).order_by("anno", "mese")

    cashflow = {}
    for r in query.all():
        chiave = f"{int(r.anno)}-{int(r.mese):02d}"
        if chiave not in cashflow:
            cashflow[chiave] = {"periodo": chiave, "entrate": 0, "uscite": 0}
        if r.direzione == "Entrata":
            cashflow[chiave]["entrate"] = round(float(r.totale), 2)
        else:
            cashflow[chiave]["uscite"] = round(abs(float(r.totale)), 2)

    return sorted(cashflow.values(), key=lambda x: x["periodo"])


# ── TRANSAZIONI ───────────────────────────────────────────────────────────────

@app.get("/transazioni")
def get_transazioni(
    anno: int = None,
    mese: int = None,
    conto_id: int = None,
    direzione: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    query = filtro_reali(db.query(Transazione))

    if anno:
        query = query.filter(extract("year", Transazione.data) == anno)
    if mese:
        query = query.filter(extract("month", Transazione.data) == mese)
    if conto_id:
        query = query.filter(Transazione.account_id == conto_id)
    if direzione:
        query = query.filter(Transazione.direzione == direzione)

    totale = query.count()
    transazioni = query.order_by(Transazione.data.desc()).offset(offset).limit(limit).all()

    return {
        "totale": totale,
        "offset": offset,
        "limit": limit,
        "transazioni": [
            {
                "id": t.id,
                "data": t.data.isoformat(),
                "direzione": t.direzione,
                "importo": float(t.importo),
                "descrizione": t.descrizione,
                "conto": t.account.nome_conto,
                "causale": t.entrata.nome_entrata if t.entrata else t.uscita.nome_uscita if t.uscita else None,
            }
            for t in transazioni
        ],
    }


# ── PER CATEGORIA ─────────────────────────────────────────────────────────────

@app.get("/dashboard/per-categoria")
def get_per_categoria(anno: int = None, mese: int = None, db: Session = Depends(get_db)):
    oggi = date.today()
    anno = anno or oggi.year
    mese = mese or oggi.month

    uscite = filtro_reali(db.query(
        Uscita.categoria,
        func.sum(Transazione.importo).label("totale")
    ).join(Transazione, Transazione.uscita_id == Uscita.id)).filter(
        extract("year", Transazione.data) == anno,
        extract("month", Transazione.data) == mese,
    ).group_by(Uscita.categoria).all()

    entrate = filtro_reali(db.query(
        Entrata.categoria,
        func.sum(Transazione.importo).label("totale")
    ).join(Transazione, Transazione.entrata_id == Entrata.id)).filter(
        extract("year", Transazione.data) == anno,
        extract("month", Transazione.data) == mese,
    ).group_by(Entrata.categoria).all()

    return {
        "uscite_per_categoria": [
            {"categoria": r.categoria, "totale": round(abs(float(r.totale)), 2)}
            for r in uscite
        ],
        "entrate_per_categoria": [
            {"categoria": r.categoria, "totale": round(float(r.totale), 2)}
            for r in entrate
        ],
    }


# ── EXPORT EXCEL ────────────────────────────────────────────────────────────

@app.get("/export/excel")
def export_excel(anno: int = None, db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws_tx = wb.active
    ws_tx.title = "Transazioni"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1B3A36")

    headers = ["Data", "Conto", "Direzione", "Causale", "Descrizione", "Importo"]
    for col, h in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    query = filtro_reali(db.query(Transazione))
    if anno:
        query = query.filter(extract("year", Transazione.data) == anno)
    transazioni = query.order_by(Transazione.data.desc()).all()

    for row, t in enumerate(transazioni, 2):
        causale = t.entrata.nome_entrata if t.entrata else t.uscita.nome_uscita if t.uscita else ""
        ws_tx.cell(row=row, column=1, value=t.data.isoformat())
        ws_tx.cell(row=row, column=2, value=t.account.nome_conto)
        ws_tx.cell(row=row, column=3, value=t.direzione)
        ws_tx.cell(row=row, column=4, value=causale)
        ws_tx.cell(row=row, column=5, value=t.descrizione)
        ws_tx.cell(row=row, column=6, value=float(t.importo))

    for col in ws_tx.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_tx.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws_conti = wb.create_sheet("Conti")
    for col, h in enumerate(["Nome Conto", "Categoria"], 1):
        cell = ws_conti.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, a in enumerate(db.query(Account).order_by(Account.categoria).all(), 2):
        ws_conti.cell(row=row, column=1, value=a.nome_conto)
        ws_conti.cell(row=row, column=2, value=a.categoria)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"patrimonio_{anno or 'tutto'}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )